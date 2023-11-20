from openpyxl import load_workbook

# Open the Excel file
workbook = load_workbook('London Underground data.xlsx')

# Select the sheet you want to work with
sheet = workbook['Sheet1']

# Selecting which data to work on
start_row = 1  # Starting row
end_row = 757   # Ending row
start_col = 1  # Column A
end_col = 4    # Column B
london_underground = []
# Iterate through the specified range of cells and remove None values from each row
for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True):
    filtered_data = [item for item in row if item is not None]
    filtered_tuple = tuple(filtered_data)
    london_underground.append(list(filtered_tuple))

# filter the nested list to only contain the stations with an adjacent station and a number of minutes.
filtered_list = [sublist for sublist in london_underground if len(sublist) == 4]
filtered_list_1 = [sublist[1:] for sublist in filtered_list]

london_underground_dict = {}  # create dictionary to store train lines and stations

for sublist in filtered_list:
    key = sublist[0]
    values = tuple(sublist[1:])
    if key in london_underground_dict:
        london_underground_dict[key] += tuple(values)  # Concatenate the existing tuple
    else:
        london_underground_dict[key] = tuple(values)  # Create a new tuple

filtered_dict = {}  # Create a new dictionary to store the train lines and stations in a weighted edge structure

for key, values_tuple in london_underground_dict.items():
    grouped_tuples = tuple(values_tuple[i:i + 3] for i in range(0, len(values_tuple), 3))
    filtered_dict[key] = grouped_tuples

# Take only the stations from the london_underground nested list
stations_all = [sublist for sublist in london_underground if len(sublist) == 2]
# make a new list containing all the stations
stations = [sublist[1] for sublist in stations_all]

# Take only the values of the filtered_dict dictionary to take only the stations, adjacent stations and the minutes
new_filtered_list_1 = [item for index, item in enumerate(filtered_list_1) if item not in stations[:index]]
# print(new_filtered_list_1) # list without train lines

bidirectional_edges = []
for edge in filtered_list_1:
    source, target, weight = edge
    bidirectional_edges.append((source, target, weight))
    bidirectional_edges.append((target, source, weight))

# Remove extra spaces from all the stations
stations_1 = [station.strip() for station in stations]
# Normalize the station names in the bidirectional_edges list
# Remove leading and trailing spaces from station names in the bidirectional_edges list
bidirectional_edges = [(edge[0].strip(), edge[1].strip(), edge[2]) for edge in bidirectional_edges]
#print(bidirectional_edges)

task_4_edge = [sublist[:2] for sublist in new_filtered_list_1]
task_4_edges = [[item.strip() for item in sublist[:2]] for sublist in new_filtered_list_1]
